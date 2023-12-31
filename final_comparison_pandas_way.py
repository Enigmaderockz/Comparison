
import openpyxl
import os
import sys
from operator import itemgetter
import concurrent.futures
import pandas as pd
import itertools
import base64
from concurrent.futures import ThreadPoolExecutor
import heapq


def convert_row_to_dict(row):
    # Convert the row from a list of tuples to a dictionary
    return {col: val for col, val in row}

def chunks(iterable, chunk_size):
    """Yield successive chunks from iterable."""
    iterator = iter(iterable)
    while True:
        chunk = list(itertools.islice(iterator, chunk_size))
        if not chunk:
            return
        yield chunk

def parallel_sort_rows(rows, sort_keys=None):
    # Number of parallel threads (adjust according to your system)
    num_threads = 4

    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        sorted_rows = list(executor.map(lambda chunk: sorted(chunk, key=lambda row: mixed_type_sort_key(row, sort_keys)), chunks(rows, chunk_size=10000)))

    return list(heapq.merge(*sorted_rows, key=lambda row: mixed_type_sort_key(row, sort_keys)))



def process_batches(sorted_rows, sort_keys, batch_size):
    result_batches = []
    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(parallel_sort_rows, batch, sort_keys) for batch in chunks(sorted_rows, 5000)]
    sorted_rows = list(heapq.merge(*[f.result() for f in futures], key=lambda row: mixed_type_sort_key(row, sort_keys)))
    return sorted_rows

def parallel_compare_rows(row1, row2):
    # Compare each row from both files and return the differences
    row_diff = False
    diff_cols = []

    for col, val1 in row1.items():
        val2 = row2[col]

        if pd.isnull(val1) and pd.isnull(val2):
            val1 = "NULL"
            val2 = "NULL"
            continue

        if val1 != val2:
            row_diff = True
            diff_cols.append(col)

    return row_diff, diff_cols, row1, row2

def read_file(file_path, column_mapping=None):
    _, file_extension = os.path.splitext(file_path)
    if file_extension == ".csv":
        for chunk in pd.read_csv(file_path, chunksize=10000, dtype=str, delimiter=","):
            if column_mapping:
                chunk = chunk.rename(columns=column_mapping)
            for row in chunk.to_dict("records"):
                yield row
    elif file_extension == ".xlsx":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            header = [cell.value for cell in sheet[1]]
            if column_mapping:
                header = [column_mapping.get(col, col) for col in header]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {header[i]: value for i, value in enumerate(row)}
                yield row_dict
    else:
        raise ValueError("Unsupported file format")


def mixed_type_sort_key(row, keys=None):
    if keys is None:
        keys = row.keys()  # Use all keys if none are specified

    sort_key = []
    for key in keys:
        value = row[key]
        if isinstance(value, str):
            sort_key.append((1, value))
        else:
            sort_key.append((0, value))
    return tuple(sort_key)

def compare_csv_files(file1, file2, outfile, sort_keys=None, exclude_keys=None, column_mapping=None):

    
    usr_exclusion_msg = (
        f"Columns to be excluded in comparison based on user input: {', '.join(exclude_keys)}"
        if exclude_keys
        else "Columns to be excluded in comparison based on user input: None"
    )

    with open(outfile, "w") as outfile:
        # Read CSV files into DataFrames
        df1 = pd.DataFrame(read_file(sys.argv[1], column_mapping))
        df2 = pd.DataFrame(read_file(sys.argv[2], column_mapping))

         # Sort DataFrames based on sort_keys
        
        
   
        # Read the first rows of both DataFrames to get headers
        first_row1 = df1.iloc[0]
        first_row2 = df2.iloc[0]

        header1 = list(first_row1.index)
        header2 = list(first_row2.index)

        # Find mismatched column headers
        mismatch_column_headers = list(set(header1) - set(header2)) + list(set(header2) - set(header1))

        col_header_diff_msg = (
            f"Column names differ based on column header mismatch between both files: {', '.join(mismatch_column_headers)}"
            if mismatch_column_headers
            else "Column names differ based on column header mismatch between both files: None"
        )

        # Adjust exclude_keys based on mismatched column headers
        exclude_keys = (
            [] if exclude_keys is None else exclude_keys
        )  # Initialize exclude_keys as an empty list if it's None
        exclude_keys.extend(mismatch_column_headers)
        

        total_col_exclusion_msg = (
            f"Total columns to be excluded in comparison: {', '.join(exclude_keys)}"
            if exclude_keys
            else "Total columns to be excluded in comparison: None"
        )

        combined_msg = f"{usr_exclusion_msg}\n{col_header_diff_msg}\n{total_col_exclusion_msg}"
        print(combined_msg)

        combined_html_msg = (
            f"<ul>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{usr_exclusion_msg}</span></li>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{col_header_diff_msg}</span></li>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{total_col_exclusion_msg}</span></li>"
            f"</ul>"
        )
            
            # Exclude specified columns
        if exclude_keys:
            df1 = df1.drop(columns=exclude_keys, errors="ignore")
            df2 = df2.drop(columns=exclude_keys, errors="ignore")

        header1 = (
            header1
            if not exclude_keys
            else [item for item in header1 if item not in exclude_keys]
        )

        sorted_rows1 = process_batches(df1.to_dict('records'), sort_keys, 5000)
        sorted_rows2 = process_batches(df2.to_dict('records'), sort_keys, 5000)

        
            
       
        
        # Initialize HTML string with a table
        html = "<html>\n<head>\n<style>\nbody {font-family: Helvetica;}\n</style>\n</head>\n<body>\n"

        # Read the contents of output.csv
        with open("output.csv", "r") as file:
            csv_content = [line.strip() for line in file if line.strip()]

        # Check if both lines contain ": 0"
        recon_status = all(": 0" in line for line in csv_content)

        # Encode the contents as Base64
        csv_base64 = base64.b64encode("".join(csv_content).encode()).decode("utf-8")

        # Embed the CSV as a data URI
        html += '<div style="font-family: Helvetica, sans-serif; margin-top: 20px;">\n'
        html += '<table style="border-collapse: collapse;">'

        # Add the recon status to the HTML
        if recon_status:
            html += '<th colspan="100" style="color: green; font-weight: bold;">Recon Status: OK</th></tr>'
        else:
            html += '<th colspan="100" style="color: red; font-weight: bold;">Recon Status: NOK</th></tr>'

        html += "<tr><td></td></tr>\n"
        html += (
            '<tr><td><a href="data:text/csv;base64,'
            + csv_base64
            + '" download="output.csv">Recon summary</a></td></tr>\n'
        )
        html += '<tr style="height: 20px;"><td></td></tr>\n'  # Add spacing of 20 pixels
        html += "</table>\n"
        html += "</div>\n"
        html += (
            '<tr><th style="border: 1px solid black;">Comparison Summary:-</th></tr>'
        )
        html += '<table style="border: 1px solid black; border-collapse: collapse;">\n'



        # Initialize counters
        num_records = len(sorted_rows1)
        num_diff_records = 0
        num_records_file1_not_in_file2 = 0
        num_records_file2_not_in_file1 = 0
        

        # Convert DataFrames to sorted tuples of rows
        # Convert DataFrames to sorted tuples of rows
        sorted_rows1_set = set(
            tuple(sorted(sorted_row.items())) for sorted_row in sorted_rows1
        )
        sorted_rows2_set = set(
            tuple(sorted(sorted_row.items())) for sorted_row in sorted_rows2
        )
        num_diff_records = len(sorted_rows1_set.symmetric_difference(sorted_rows2_set))
        num_records_file1_not_in_file2 = len(
            sorted_rows1_set.difference(sorted_rows2_set)
        )
        num_records_file2_not_in_file1 = len(
            sorted_rows2_set.difference(sorted_rows1_set)
        )

        # Generate HTML for comparison summary
        summary_html = '<div style="font-size: 15px;">'
        summary_html += f"<p>Total Records in each file: {num_records}</p>"
        summary_html += (
            f"<p>Number of Rows with Differences: {int(num_diff_records)}</p>"
        )
        summary_html += f"<p>Number of records from {file1} differ from {file2}:  {num_records_file1_not_in_file2}</p>"
        summary_html += f"<p>Number of records from {file2} differ from {file1}:  {num_records_file2_not_in_file1}</p>"
        summary_html += "</div>\n"

        # Add the summary HTML to the start of the HTML string
        html += summary_html

        # Check if there are any differences in the data and generate HTML table for differences
        if num_diff_records == 0:
            print(
                f"{file1} and {file2} have no differences. Please check the generated HTML for more details."
            )
        elif num_diff_records > 0:
            print(
                f"{file1} and {file2} have differences by {num_diff_records}. Please check the generated HTML for more details."
            )

            diff_rows = []
            diff_columns = set()

            # Compare each row from both files and add different rows to diff_rows list
            with concurrent.futures.ThreadPoolExecutor() as executor:
                futures = []
                for row1, row2 in zip(sorted_rows1, sorted_rows2):
                    future = executor.submit(parallel_compare_rows, row1, row2)
                    futures.append(future)

                diff_rows = []
                diff_columns = set()

                for row_num, future in enumerate(concurrent.futures.as_completed(futures), start=1):
                    row_diff, diff_cols, row1, row2 = future.result()

                    if row_diff:
                        diff_rows.append((row_num, diff_cols, row1, row2))
                        print(diff_rows)
                        diff_columns.update(diff_cols)
                    
                    # Calculate and print the progress in percentage
                    progress = row_num / num_records * 100
                    print(f"\rComparison progress: {progress:.2f}%", end="")

            # Print the column names with differences below the HTML table
            html += (
                '<div style="font-family: Helvetica, sans-serif; margin-top: 20px;">\n'
            )
            html += '<table style="border-collapse: collapse;">'
            html += '<tr style="background-color: #92b9bf; border: 1px solid black;"><th colspan="100" style="border: 1px solid black;">Columns with Differences:-</th></tr>'
            for col in diff_columns:
                html += f"<tr><td>{col}</td></tr>"
            html += "</table>\n"
            html += "</div>\n"

            # Write the HTML table for differences to the output file
            # Write the HTML table for differences to the output file
            html += '<table style="font-family: Helvetica, sans-serif; font-size:14px;border-collapse: collapse; width: 100%; border: 1px solid black; border-spacing: 0px;">\n'
            html += '<tr style="background-color: #a0d1dd; border: 1px solid black;"><th colspan="100" style="border: 1px solid black;">Differences</th></tr>'
            html += '<tr style="background-color: #a0d1dd; border: 1px solid black;"><th style="border: 1px solid black;">File</th><th style="border: 1px solid black;">Row Number</th>'

            # Write the headers of CSV files
            for col in header1:
                html += f'<th style="border: 1px solid black;">{col}</th>'

            html += "</tr>\n"
            # Write the rows with differences to the HTML file

            for diff_row in diff_rows:
                for i in range(2):  # Iterate twice for both files
                    file_data = [(file1, diff_row[2]), (file2, diff_row[3])][i]
                    file_name, row_data = file_data

                    html += '<tr style="border: 1px solid black;">'

                    # Write the file name and row number
                    html += f'<td style="border: 1px solid black;">{file_name}</td>'
                    html += f'<td style="border: 1px solid black;">{diff_row[0]}</td>'

                    # Write the values of each column of the row
                    for col_idx, col in enumerate(header1):
                        cell_style = ''
                        cell_value = row_data[col]  # Access value by column name

                        if col in diff_row[1]:
                            cell_style = 'background-color: #ffcfbf;font-weight: bold;'

                        if pd.isnull(cell_value):
                            cell_value = 'NULL'

                        html += f'<td style="border: 1px solid black; {cell_style}">{cell_value}</td>'

                    html += '</tr>\n'
            # Close the HTML table
            html += "</table>\n"
        # Write the contents of the entire HTML string to the output file
        outfile.write(html)
        


# Extract sort_keys from system arguments
# Extract sort_keys from system arguments
column_mapping = None
sort_keys = None
exclude_keys = None
for arg in sys.argv:
    if "include=" in arg:
        _, value = arg.split("=")
        if value.lower() != "none":
            if value == "*":
                # Read the headers from the first file
                with open(sys.argv[1], "r") as f:
                    rows = read_file(sys.argv[1])
                    first_row = next(rows)
                    sort_keys = list(first_row.keys())
            else:
                sort_keys = value.split(";")
    elif "exclude=" in arg:
        _, value = arg.split("=")
        exclude_keys = value.split(";")
    elif arg.startswith("column_mapping="):
        _, value = arg.split("=")
        if value.lower() == "y":
            with open("mapping.txt", "r") as f:
                column_mapping = {}
                lines = f.readlines()
                for line in lines:
                    mapping = line.strip().split(":")
                    if len(mapping) == 2:
                        src_cols = mapping[0].split(",")
                        dest_cols = mapping[1].split(",")
                        for src_col, dest_col in zip(src_cols, dest_cols):
                            column_mapping[dest_col] = src_col
        else:
            column_mapping = None


compare_csv_files(
sys.argv[1],
sys.argv[2],
sys.argv[3],
sort_keys=sort_keys,
exclude_keys=exclude_keys,
column_mapping=column_mapping,
)



