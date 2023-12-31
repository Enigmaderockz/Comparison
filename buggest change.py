find.py

import csv
import shutil
import sys
import os
import pandas as pd
import comp2


def fetch_delimiter(file_path):
    _, file_extension = os.path.splitext(file_path)
    if file_extension == ".csv" or file_extension == ".dat":
        with open(file_path, "r") as file:
            dialect = csv.Sniffer().sniff(file.read(102400))
            return dialect.delimiter


def read_mapping(file):
    with open(file, "r") as f:
        mapping = {}
        for line in f:
            cols = line.strip().split(":")
            source_cols = cols[0].split(",")
            target_cols = cols[1].split(",")
            for i in range(len(source_cols)):
                mapping[source_cols[i]] = target_cols[i]
    return mapping


def update_column_names(file, column_mapping):
    with open(file, "r") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = list(reader)
        new_fieldnames = [column_mapping.get(name, name) for name in reader.fieldnames]

    with open(file, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=new_fieldnames, delimiter=delimiter)
        writer.writeheader()
        for row in rows:
            new_row = {
                column_mapping.get(name, name): value for name, value in row.items()
            }
            writer.writerow(new_row)


def create_temp_files(file1, file2, column_mapping=None, buffer_size=1024 * 1024):
    global delimiter
    delimiter = fetch_delimiter(file1)
    print(delimiter)
    # Copy file1
    with open(file1, "rb") as fsrc, open("a_tmp.csv", "wb") as fdst:
        shutil.copyfileobj(fsrc, fdst, buffer_size)

    # Copy file2
    with open(file2, "rb") as fsrc, open("b_tmp.csv", "wb") as fdst:
        shutil.copyfileobj(fsrc, fdst, buffer_size)

    if column_mapping is not None:
        column_mapping = read_mapping("mapping.txt")
        print(column_mapping)
        update_column_names("a_tmp.csv", column_mapping)
        update_column_names("b_tmp.csv", column_mapping)


def perform_recon_on_files(file1, file2, col_check=None, column_mapping=None):
    print(col_check)
    delimiter1 = ","
    delimiter2 = "|"
    if col_check is not None:
        # Read the CSV files
        df1 = pd.read_csv(file1, dtype=str)
        df2 = pd.read_csv(file2, dtype=str)

        col_check_list = col_check.split(",")

        if column_mapping is not None:
            col_check_list = [column_mapping.get(col, col) for col in col_check_list]

        # Create sets for faster lookup
        set1 = set(tuple(row) for row in df1[col_check_list].values)
        set2 = set(tuple(row) for row in df2[col_check_list].values)

        # Find extra records
        extra_records1 = df1[~df1[col_check_list].apply(tuple, axis=1).isin(set2)]
        extra_records2 = df2[~df2[col_check_list].apply(tuple, axis=1).isin(set1)]

        # Write the results to output.csv
        with open("output1.csv", "w", newline="") as output_file:
            writer = csv.writer(
                output_file, delimiter=delimiter1
            )  # Use ',' as delimiter

            # Write the first set of extra records
            # writer.writerow([f"Extra records in {file1} which are not present in {file2} based on " + col_check + ": " + str(len(extra_records1))])
            # writer.writerow([])
            if not extra_records1.empty:
                writer.writerow(extra_records1.columns)
                extra_records1 = extra_records1.fillna("")
                writer.writerows(extra_records1.values)
            else:
                writer.writerow([])

            writer.writerow([])

        with open("output2.csv", "w", newline="") as output_file:
            writer = csv.writer(output_file, delimiter=delimiter1)

            # Write the second set of extra records
            # writer.writerow([f"Extra records in {file2} which are not present in {file1} based on " + col_check + ": " + str(len(extra_records2))])
            # writer.writerow([])
            if not extra_records2.empty:
                writer.writerow(extra_records2.columns)
                extra_records2 = extra_records2.fillna("")
                writer.writerows(extra_records2.values)
            else:
                writer.writerow([])

        # Update the temporary files
        df1 = df1[df1[col_check_list].apply(tuple, axis=1).isin(set2)]
        df2 = df2[df2[col_check_list].apply(tuple, axis=1).isin(set1)]
        df1.to_csv(file1, index=False, sep=delimiter1)
        df2.to_csv(file2, index=False, sep=delimiter2)
        return len(extra_records1), len(extra_records2)


# Run the compare_csv function with the required parameters
file1 = "a.dat"
file2 = "b.dat"
column_mapping = None
col_check = None

for arg in sys.argv:
    if arg.startswith("col_check="):
        _, value = arg.split("=")
        if value.lower() != "none":
            if value == "*":
                with open(file2, "r") as f:
                    reader = csv.reader(f)
                    col_check = next(reader)
                    col_check = ",".join(col_check)
                    print(col_check)
            else:
                col_check = value
    elif arg.startswith("column_mapping="):
        _, value = arg.split("=")
        if value.lower() == "y":
            with open("mapping.txt", "r") as f:
                column_mapping = {}
                lines = f.readlines()
                for line in lines:
                    mapping = line.strip().split(":")
                    if len(mapping) == 2:
                        src_cols = mapping[0].split(",")
                        dest_cols = mapping[1].split(",")
                        for src_col, dest_col in zip(src_cols, dest_cols):
                            column_mapping[src_col] = dest_col
        else:
            column_mapping = None


def recon(file1, file2, col_check=col_check, column_mapping=column_mapping):
    create_temp_files(file1, file2, column_mapping)
    rec1, rec2 = perform_recon_on_files(
        "a_tmp.csv", "b_tmp.csv", col_check=col_check, column_mapping=column_mapping
    )
    print(rec1)
    comp2.compare_csv_files(
        "a_tmp.csv", "b_tmp.csv", "output.html", rec1, rec2, col_check=col_check
    )


recon(file1, file2, col_check=col_check, column_mapping=column_mapping)

comp2.py

import csv
import openpyxl
import os
import sys
from operator import itemgetter
import concurrent.futures
import pandas as pd
import itertools
import base64


def read_file(file_path, column_mapping=None):
    _, file_extension = os.path.splitext(file_path)
    if file_extension == ".csv":
        for chunk in pd.read_csv(file_path, chunksize=10000, dtype=str, delimiter=","):
            if column_mapping:
                chunk = chunk.rename(columns=column_mapping)
            for row in chunk.to_dict("records"):
                yield row
    elif file_extension == ".xlsx":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            header = [cell.value for cell in sheet[1]]
            if column_mapping:
                header = [column_mapping.get(col, col) for col in header]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {header[i]: value for i, value in enumerate(row)}
                yield row_dict
    else:
        raise ValueError("Unsupported file format")


def mixed_type_sort_key(row, keys=None):
    if keys is None:
        keys = row.keys()  # Use all keys if none are specified

    sort_key = []
    for key in keys:
        value = row[key]
        if isinstance(value, str):
            sort_key.append((1, value))
        else:
            sort_key.append((0, value))
    return tuple(sort_key)


def compare_csv_files(
    file1,
    file2,
    outfile,
    rec1,
    rec2,
    sort_keys=None,
    exclude_keys=None,
    column_mapping=None,
    col_check=None,
):
    print(column_mapping)
    usr_exclusion_msg = (
        f"Columns to be excluded in comparison based on user input: {', '.join(exclude_keys)}"
        if exclude_keys
        else "Columns to be excluded in comparison based on user input: None"
    )

    # Initialize counters
    num_records = 0
    num_diff_records = 0
    num_records_file1_not_in_file2 = 0
    num_records_file2_not_in_file1 = 0

    # Open the input files and output file for writing HTML table
    with open(file1, "r") as f1, open(file2, "r") as f2, open(outfile, "w") as outfile:
        rows1 = read_file(file1, column_mapping)
        rows2 = read_file(file2, column_mapping)

        try:
            first_row1 = next(rows1)
            first_row2 = next(rows2)
        except StopIteration:
            raise ValueError("One or both files are empty")

        rows1 = itertools.chain([first_row1], rows1)
        rows2 = itertools.chain([first_row2], rows2)

        header1 = list(first_row1.keys())
        header2 = list(first_row2.keys())

        mismatch_column_headers = list(set(header1) - set(header2)) + list(
            (set(header2) - set(header1))
        )

        col_header_diff_msg = (
            f"Column names differ based on column header mismatch between both files: {', '.join(mismatch_column_headers)}"
            if mismatch_column_headers
            else "Column names differ based on column header mismatch between both files: None"
        )

        exclude_keys = list(
            set(
                [] + mismatch_column_headers
                if exclude_keys is None
                else exclude_keys + mismatch_column_headers
            )
        )
        total_col_exclusion_msg = (
            f"Total columns to be excluded in comparison: {', '.join(exclude_keys)}"
            if exclude_keys
            else "Total columns to be excluded in comparison: None"
        )

        combined_msg = (
            f"{usr_exclusion_msg}\n{col_header_diff_msg}\n{total_col_exclusion_msg}"
        )
        print(combined_msg)

        combined_html_msg = (
            f"<ul>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{usr_exclusion_msg}</span></li>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{col_header_diff_msg}</span></li>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{total_col_exclusion_msg}</span></li>"
            f"</ul>"
        )

        # Exclude the specified columns
        if len(exclude_keys) > 0:
            rows1 = [
                {k: v for k, v in row.items() if k not in exclude_keys} for row in rows1
            ]
            rows2 = [
                {k: v for k, v in row.items() if k not in exclude_keys} for row in rows2
            ]

        # Check if both files have same headers/fieldnames
        if header1 != header2:
            # raise ValueError("Headers not matching in both files")
            pass

        header1 = (
            header1
            if not exclude_keys
            else [item for item in header1 if item not in exclude_keys]
        )
        print(header1)

        sort_key_func = lambda row: mixed_type_sort_key(row, sort_keys)

        # Use a ThreadPoolExecutor to sort the rows in parallel
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future_to_rows = {
                executor.submit(sorted, rows, key=sort_key_func): rows
                for rows in [rows1, rows2]
            }
            sorted_rows1, sorted_rows2 = [
                future.result()
                for future in concurrent.futures.as_completed(future_to_rows)
            ]

        # Count the number of records in each file
        num_records = len(sorted_rows1)
        print(num_records)

        html = "<html>\n<head>\n<style>\nbody {font-family: Helvetica;}\n</style>\n</head>\n<body>\n"

        # Read the contents of output.csv
        with open("output1.csv", "r") as file:
            csv_content1 = file.read()

        with open("output2.csv", "r") as file:
            csv_content2 = file.read()

        # Encode the contents as Base64
        csv_base64_1 = base64.b64encode("".join(csv_content1).encode()).decode("utf-8")
        csv_base64_2 = base64.b64encode("".join(csv_content2).encode()).decode("utf-8")

        # Embed the CSV as a data URI
        html += '<div style="font-family: Helvetica, sans-serif; margin-top: 20px;">\n'
        html += '<table style="border-collapse: collapse;">'
        html += "<tr><td></td></tr>\n"

        if rec1 == 0 and rec2 == 0:
            html += '<th colspan="100" style="color: green; font-weight: bold;">Recon Status: OK</th></tr>'
            html += f"<tr><td>Extra records in {file1} which are not present in {file2} based on columns {col_check}: {str(rec1)}</td></tr>\n"
            html += f"<tr><td>Extra records in {file2} which are not present in {file1} based on columns {col_check}: {str(rec2)}</td></tr>\n"
        else:
            html += '<th colspan="100" style="color: red; font-weight: bold;">Recon Status: NOK</th></tr>'

            if rec2 != 0 and rec1 != 0:
                html += f"<tr><td>Extra records in {file1} which are not present in {file2} based on columns {col_check}: {str(rec1)}</td></tr>\n"
                html += (
                    '<tr><td><a href="data:text/csv;base64,'
                    + csv_base64_1
                    + '" download="output.csv">Recon summary1</a></td></tr>\n'
                )
                html += f"<tr><td>Extra records in {file2} which are not present in {file1} based on columns {col_check}: {str(rec2)}</td></tr>\n"
                html += (
                    '<tr><td><a href="data:text/csv;base64,'
                    + csv_base64_2
                    + '" download="output.csv">Recon summary2</a></td></tr>\n'
                )

            if rec1 != 0 and rec2 == 0:
                html += f"<tr><td>Extra records in {file1} which are not present in {file2} based on columns {col_check}: {str(rec1)}</td></tr>\n"
                html += (
                    '<tr><td><a href="data:text/csv;base64,'
                    + csv_base64_1
                    + '" download="output.csv">Recon summary1</a></td></tr>\n'
                )
                html += f"<tr><td>Extra records in {file2} which are not present in {file1} based on columns {col_check}: {str(rec2)}</td></tr>\n"

            if rec2 != 0 and rec1 == 0:
                html += f"<tr><td>Extra records in {file2} which are not present in {file1} based on columns {col_check}: {str(rec2)}</td></tr>\n"
                html += (
                    '<tr><td><a href="data:text/csv;base64,'
                    + csv_base64_2
                    + '" download="output.csv">Recon summary2</a></td></tr>\n'
                )
                html += f"<tr><td>Extra records in {file1} which are not present in {file2} based on columns {col_check}: {str(rec1)}</td></tr>\n"

        html += '<tr style="height: 20px;"><td></td></tr>\n'  # Add spacing of 20 pixels
        html += "</table>\n"
        html += "</div>\n"

        html += (
            '<tr><th style="border: 1px solid black;">Comparison Summary:-</th></tr>'
        )
        html += '<table style="border: 1px solid black; border-collapse: collapse;">\n'

        # Compare each row from both files

        sorted_rows1_set = set(
            tuple(sorted(sorted_row.items())) for sorted_row in sorted_rows1
        )
        sorted_rows2_set = set(
            tuple(sorted(sorted_row.items())) for sorted_row in sorted_rows2
        )
        num_diff_records = len(sorted_rows1_set.symmetric_difference(sorted_rows2_set))
        num_records_file1_not_in_file2 = len(
            sorted_rows1_set.difference(sorted_rows2_set)
        )
        num_records_file2_not_in_file1 = len(
            sorted_rows2_set.difference(sorted_rows1_set)
        )

        # col = "/v/g/"

        # Generate HTML for comparison summary
        summary_html = '<div style="font-size: 15px;">'
        summary_html += f"<p>Total Records in each file: {num_records}</p>"
        summary_html += (
            f"<p>Number of Rows with Differences: {int(num_diff_records)}</p>"
        )
        summary_html += f"<p>Number of records from {file1} differ from {file2}:  {num_records_file1_not_in_file2}</p>"
        summary_html += f"<p>Number of records from {file2} differ from {file1}:  {num_records_file2_not_in_file1}</p>"
        # summary_html += f"<p>Number of records from {file1.replace('_tmp', '')} differ from {col}.{file2.replace('_tmp', '')}:  {num_records_file1_not_in_file2}</p>"
        # summary_html += f"<p>Number of records from {file2.replace('_tmp', '')} differ from {col}.{file1.replace('_tmp', '')}:  {num_records_file2_not_in_file1}</p>"
        summary_html += f"<p>Please Note: {combined_html_msg}</p>"
        summary_html += "</div>\n"

        # Add the summary HTML to the start of the HTML string
        html += summary_html

        # Check if there are any differences in the data and generate HTML table for differences
        if num_diff_records == 0:
            print(
                f"{file1} and {file2} have no differences. Please check the generated HTML for more details."
            )
        elif num_diff_records > 0:
            print(
                f"{file1} and {file2} have differences by {num_diff_records}. Please check the generated HTML for more details."
            )

            diff_rows = []
            diff_columns = set()

            # Compare each row from both files and add different rows to diff_rows list
            for row_num, (row1, row2) in enumerate(
                zip(sorted_rows1, sorted_rows2), start=1
            ):
                row_diff = False
                diff_cols = []

                # Check if the values in each column of the row are the same
                for col, val1 in row1.items():
                    val2 = row2[col]

                    if pd.isnull(val1) and pd.isnull(val2):
                        val1 = "NULL"
                        val2 = "NULL"
                        continue

                    # If the values in the columns are different, add the column index to diff_cols list
                    if val1 != val2:
                        row_diff = True
                        diff_cols.append(col)
                        diff_columns.add(col)

                # If the row has differences, add it to the diff_rows list along with the indices of different columns
                if row_diff:
                    diff_rows.append((row_num, diff_cols, row1, row2))

            # Print the column names with differences below the HTML table
            html += (
                '<div style="font-family: Helvetica, sans-serif; margin-top: 20px;">\n'
            )
            html += '<table style="border-collapse: collapse;">'
            html += '<tr style="background-color: #92b9bf; border: 1px solid black;"><th colspan="100" style="border: 1px solid black;">Columns with Differences:-</th></tr>'
            for col in diff_columns:
                html += f"<tr><td>{col}</td></tr>"
            html += "</table>\n"
            html += "</div>\n"

            # Write the HTML table for differences to the output file
            html += '<table style="font-family: Helvetica, sans-serif; font-size:14px;border-collapse: collapse; width: 100%; border: 1px solid black; border-spacing: 0px;">\n'
            html += '<tr style="background-color: #a0d1dd; border: 1px solid black;"><th colspan="100" style="border: 1px solid black;">Differences</th></tr>'
            html += '<tr style="background-color: #a0d1dd; border: 1px solid black;"><th style="border: 1px solid black;">File</th><th style="border: 1px solid black;">Row Number</th>'

            # Write the headers of CSV files
            for col in header1:
                html += f'<th style="border: 1px solid black;">{col}</th>'

            html += "</tr>\n"

            # Write the rows with differences to the HTML file
            for diff_row in diff_rows:
                html += '<tr style="border: 1px solid black;">'
                html += f'<td style="border: 1px solid black;">{file1}</td>'
                html += f'<td style="border: 1px solid black;">{diff_row[0]}</td>'

                # Write the values of each column of the row
                for col in header1:
                    if col in diff_row[2]:
                        if col in diff_row[1]:
                            if pd.isnull(diff_row[2][col]):
                                html += f'<td style="border: 1px solid black; background-color: #ffcfbf;font-weight: bold;">NULL</td>'
                            else:
                                html += f'<td style="border: 1px solid black; background-color: #ffcfbf;font-weight: bold;">{diff_row[2][col]}</td>'
                        else:
                            if pd.isnull(diff_row[2][col]):
                                html += (
                                    f'<td style="border: 1px solid black;">NULL</td>'
                                )
                            else:
                                html += f'<td style="border: 1px solid black;">{diff_row[2][col]}</td>'
                    else:
                        html += f'<td style="border: 1px solid black;"></td>'

                html += "</tr>\n"

                html += '<tr style="border: 1px solid black;">'
                html += f'<td style="border: 1px solid black;">{file2}</td>'
                html += f'<td style="border: 1px solid black;">{diff_row[0]}</td>'

                # Write the values of each column of the row
                for col in header1:
                    if col in diff_row[3]:
                        if col in diff_row[1]:
                            if pd.isnull(diff_row[3][col]):
                                html += f'<td style="border: 1px solid black; background-color: #ffcfbf;font-weight: bold;">NULL</td>'
                            else:
                                html += f'<td style="border: 1px solid black; background-color: #ffcfbf;font-weight: bold;">{diff_row[3][col]}</td>'
                        else:
                            if pd.isnull(diff_row[3][col]):
                                html += (
                                    f'<td style="border: 1px solid black;">NULL</td>'
                                )
                            else:
                                html += f'<td style="border: 1px solid black;">{diff_row[3][col]}</td>'
                    else:
                        html += f'<td style="border: 1px solid black;"></td>'

                html += "</tr>\n"

            # Close the HTML table
            html += "</table>\n"

        # Write the contents of the entire HTML string to the output file
        outfile.write(html)


# Extract sort_keys from system arguments
column_mapping = None
sort_keys = None
exclude_keys = None
for arg in sys.argv:
    if "include=" in arg:
        _, value = arg.split("=")
        if value.lower() != "none":
            if value == "*":
                # Read the headers from the first file
                with open(sys.argv[1], "r") as f:
                    rows = read_file(sys.argv[1])
                    first_row = next(rows)
                    sort_keys = list(first_row.keys())
            else:
                sort_keys = value.split(";")
    elif "exclude=" in arg:
        _, value = arg.split("=")
        exclude_keys = value.split(";")
    elif arg.startswith("column_mapping="):
        _, value = arg.split("=")
        if value.lower() == "y":
            with open("mapping.txt", "r") as f:
                column_mapping = {}
                lines = f.readlines()
                for line in lines:
                    mapping = line.strip().split(":")
                    if len(mapping) == 2:
                        src_cols = mapping[0].split(",")
                        dest_cols = mapping[1].split(",")
                        for src_col, dest_col in zip(src_cols, dest_cols):
                            column_mapping[dest_col] = src_col
        else:
            column_mapping = None

"""
compare_csv_files(
    sys.argv[1],
    sys.argv[2],
    sys.argv[3],
    sort_keys=sort_keys,
    exclude_keys=exclude_keys,
    column_mapping=column_mapping,
)
"""


"""
Usage:

python comp1.py a.csv b.csv output.html include=*
python comp1.py a.csv b.csv output.html include=none
python comp1.py a.csv b.csv output.html include="Age;Gender;Last Name"
python comp1.py a.xlsx b.xlsx output.html include="First Name;Last Name" exclude="Gender;Country"
python comp2.py a.csv b.csv b.html column_mapping=y
"""
