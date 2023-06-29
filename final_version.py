import csv
import openpyxl
import os
import sys
from operator import itemgetter
import concurrent.futures


def read_file(file_path):
    _, file_extension = os.path.splitext(file_path)
    if file_extension == ".csv":
        with open(file_path, "r") as f:
            reader = csv.DictReader(f)
            rows = [row for row in reader]
    elif file_extension == ".xlsx":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        rows = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            header = [cell.value for cell in sheet[1]]
            sheet_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {header[i]: value for i, value in enumerate(row)}
                sheet_rows.append(row_dict)
            rows.extend(sheet_rows)
    else:
        raise ValueError("Unsupported file format")
    return rows


def mixed_type_sort_key(row, keys=None):
    if keys is None:
        keys = row.keys()  # Use all keys if none are specified

    sort_key = []
    for key in keys:
        value = row[key]
        if isinstance(value, str):
            sort_key.append((1, value))
        else:
            sort_key.append((0, value))
    return tuple(sort_key)


def compare_csv_files(file1, file2, outfile, sort_keys=None, exclude_keys=None):
    usr_exclusion_msg = (
        f"Columns to be excluded in comparison based on user input: {', '.join(exclude_keys)}"
        if exclude_keys
        else "Columns to be excluded in comparison based on user input: None"
    )

    # Initialize counters
    num_records = 0
    num_diff_records = 0
    num_records_file1_not_in_file2 = 0
    num_records_file2_not_in_file1 = 0

    # Open the input files and output file for writing HTML table
    with open(file1, "r") as f1, open(file2, "r") as f2, open(outfile, "w") as outfile:
        rows1 = read_file(file1)
        rows2 = read_file(file2)

        header1 = list(rows1[0].keys())
        header2 = list(rows2[0].keys())
        mismatch_column_headers = list(set(header1) - set(header2)) + list(
            (set(header2) - set(header1))
        )
        col_header_diff_msg = (
            f"Column names differ based on column header mismatch between both files: {', '.join(mismatch_column_headers)}"
            if mismatch_column_headers
            else "Column names differ based on column header mismatch between both files: None"
        )

        exclude_keys = list(
            set(
                [] + mismatch_column_headers
                if exclude_keys is None
                else exclude_keys + mismatch_column_headers
            )
        )
        total_col_exclusion_msg = (
            f"Total columns to be excluded in comparison: {', '.join(exclude_keys)}"
            if exclude_keys
            else "Total columns to be excluded in comparison: None"
        )

        combined_msg = (
            f"{usr_exclusion_msg}\n{col_header_diff_msg}\n{total_col_exclusion_msg}"
        )
        print(combined_msg)

        combined_html_msg = (
            f"<ul>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{usr_exclusion_msg}</span></li>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{col_header_diff_msg}</span></li>"
            f"<li><span style='color: orange; font-size: 14px; font-family: Arial; font-weight: bold;'>{total_col_exclusion_msg}</span></li>"
            f"</ul>"
        )

        # Exclude the specified columns
        if len(exclude_keys) > 0:
            rows1 = [
                {k: v for k, v in row.items() if k not in exclude_keys} for row in rows1
            ]
            rows2 = [
                {k: v for k, v in row.items() if k not in exclude_keys} for row in rows2
            ]

        header1 = list(rows1[0].keys())
        header2 = list(rows2[0].keys())

        # Check if both files have same headers/fieldnames
        if header1 != header2:
            raise ValueError("Headers not matching in both files")

        sort_key_func = lambda row: mixed_type_sort_key(row, sort_keys)

        # Use a ThreadPoolExecutor to sort the rows in parallel
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future_to_rows = {
                executor.submit(sorted, rows, key=sort_key_func): rows
                for rows in [rows1, rows2]
            }
            sorted_rows1, sorted_rows2 = [
                future.result()
                for future in concurrent.futures.as_completed(future_to_rows)
            ]

        # Count the number of records in each file
        num_records = len(sorted_rows1)

        # Initialize HTML string with a table
        html = "<html>\n<head>\n<style>\nbody {font-family: Helvetica;}\n</style>\n</head>\n<body>\n"
        html += (
            '<tr><th style="border: 1px solid black;">Comparison Summary:-</th></tr>'
        )
        html += '<table style="border: 1px solid black; border-collapse: collapse;">\n'

        # Compare each row from both files
        for row_num, (row1, row2) in enumerate(
            zip(sorted_rows1, sorted_rows2), start=1
        ):
            if row1 != row2:
                num_diff_records += 1
                if row1 not in sorted_rows2:
                    num_records_file1_not_in_file2 += 1
                if row2 not in sorted_rows1:
                    num_records_file2_not_in_file1 += 1

        # Generate HTML for comparison summary
        summary_html = '<div style="font-size: 15px;">'
        summary_html += f"<p>Total Records in each file: {num_records}</p>"
        summary_html += (
            f"<p>Number of Rows with Differences: {int(num_diff_records)}</p>"
        )
        summary_html += f"<p>Number of records from {file1} differ from {file2}:  {num_records_file1_not_in_file2}</p>"
        summary_html += f"<p>Number of records from {file2} differ from {file1}:  {num_records_file2_not_in_file1}</p>"
        summary_html += f"<p>Please Note: {combined_html_msg}</p>"
        summary_html += "</div>\n"

        # Add the summary HTML to the start of the HTML string
        html += summary_html

        # Check if there are any differences in the data and generate HTML table for differences
        if num_diff_records == 0:
            print(
                f"{file1} and {file2} have no differences. Please check the generated HTML for more details."
            )
        elif num_diff_records > 0:
            print(
                f"{file1} and {file2} have differences by {num_diff_records}. Please check the generated HTML for more details."
            )

            diff_rows = []

            # Compare each row from both files and add different rows to diff_rows list
            for row_num, (row1, row2) in enumerate(
                zip(sorted_rows1, sorted_rows2), start=1
            ):
                row_diff = False
                diff_cols = []

                # Check if the values in each column of the row are the same
                for col, val1 in row1.items():
                    val2 = row2[col]

                    # If the values in the columns are different, add the column index to diff_cols list
                    if val1 != val2:
                        row_diff = True
                        diff_cols.append(col)

                # If the row has differences, add it to the diff_rows list along with the indices of different columns
                if row_diff:
                    diff_rows.append((row_num, diff_cols, row1, row2))

            # Write the HTML table for differences to the output file
            html += '<table style="font-family: Helvetica, sans-serif; font-size:14px;border-collapse: collapse; width: 100%; border: 1px solid black; border-spacing: 0px;">\n'
            html += '<tr style="background-color: #a0d1dd; border: 1px solid black;"><th colspan="100" style="border: 1px solid black;">Differences</th></tr>'
            html += '<tr style="background-color: #a0d1dd; border: 1px solid black;"><th style="border: 1px solid black;">File</th><th style="border: 1px solid black;">Row Number</th>'

            # Write the headers of CSV files
            for col in header1:
                html += f'<th style="border: 1px solid black;">{col}</th>'

            html += "</tr>\n"

            # Write the rows with differences to the HTML file
            for diff_row in diff_rows:
                html += '<tr style="border: 1px solid black;">'
                html += f'<td style="border: 1px solid black;">{file1}</td>'
                html += f'<td style="border: 1px solid black;">{diff_row[0]}</td>'

                # Write the values of each column of the row
                for col, val in diff_row[2].items():
                    if col in diff_row[1]:
                        html += f'<td style="border: 1px solid black; background-color: #ffcfbf;font-weight: bold;">{val}</td>'
                    else:
                        html += f'<td style="border: 1px solid black;">{val}</td>'

                html += "</tr>\n"

                html += '<tr style="border: 1px solid black;">'
                html += f'<td style="border: 1px solid black;">{file2}</td>'
                html += f'<td style="border: 1px solid black;">{diff_row[0]}</td>'

                # Write the values of each column of the row
                for col, val in diff_row[3].items():
                    if col in diff_row[1]:
                        html += f'<td style="border: 1px solid black; background-color: #ffcfbf;font-weight: bold;">{val}</td>'
                    else:
                        html += f'<td style="border: 1px solid black;">{val}</td>'
                html += "</tr>\n"

            # Close the HTML table
            html += "</table>\n"

        # Write the contents of the entire HTML string to the output file
        outfile.write(html)


# Extract sort_keys from system arguments
sort_keys = None
exclude_keys = None
for arg in sys.argv:
    if "include=" in arg:
        _, value = arg.split("=")
        if value.lower() != "none":
            if value == "*":
                # Read the headers from the first file
                with open(sys.argv[1], "r") as f:
                    rows = read_file(sys.argv[1])
                    sort_keys = list(rows[0].keys())
            else:
                sort_keys = value.split(";")
    elif "exclude=" in arg:
        _, value = arg.split("=")
        exclude_keys = value.split(";")


compare_csv_files(
    sys.argv[1],
    sys.argv[2],
    sys.argv[3],
    sort_keys=sort_keys,
    exclude_keys=exclude_keys,
)

"""
Usage:

python comp1.py a.csv b.csv output.html include=*
python comp1.py a.csv b.csv output.html include=none
python comp1.py a.csv b.csv output.html include="Age;Gender;Last Name"
python comp1.py a.xlsx b.xlsx output.html include="First Name;Last Name" exclude="Gender;Country"
"""
