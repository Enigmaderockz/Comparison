import csv
import openpyxl
import os


def read_file(file_path):
    _, file_extension = os.path.splitext(file_path)
    if file_extension == ".csv":
        with open(file_path, "r") as f:
            reader = csv.DictReader(f)
            rows = [row for row in reader]
    elif file_extension == ".xlsx":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        rows = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            header = [cell.value for cell in sheet[1]]
            sheet_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {header[i]: value for i, value in enumerate(row)}
                sheet_rows.append(row_dict)
            rows.extend(sheet_rows)
    else:
        raise ValueError("Unsupported file format")
    return rows


def compare_csv_files(file1, file2, outfile):
    # Initialize counters
    num_records = 0
    num_diff_records = 0
    num_records_file1_not_in_file2 = 0
    num_records_file2_not_in_file1 = 0

    # Open the input files and output file for writing HTML table
    with open(file1, "r") as f1, open(file2, "r") as f2, open(outfile, "w") as outfile:
        rows1 = read_file(file1)
        rows2 = read_file(file2)

        header1 = list(rows1[0].keys())
        header2 = list(rows2[0].keys())

        # Check if both files have same headers/fieldnames
        if header1 != header2:
            raise ValueError("Headers not matching in both files")

        # Convert each row to a dictionary and sort the rows by the values
        sorted_rows1 = sorted([row for row in rows1], key=lambda x: list(x.values()))
        sorted_rows2 = sorted([row for row in rows2], key=lambda x: list(x.values()))

        # Count the number of records in each file
        num_records = len(sorted_rows1)

        # Initialize HTML string with a table
        html = "<html>\n<head>\n<style>\nbody {font-family: Helvetica;}\n</style>\n</head>\n<body>\n"
        html += (
            '<tr><th style="border: 1px solid black;">Comparison Summary:-</th></tr>'
        )
        html += '<table style="border: 1px solid black; border-collapse: collapse;">\n'

        # Compare each row from both files
        for row_num, (row1, row2) in enumerate(
            zip(sorted_rows1, sorted_rows2), start=1
        ):
            if row1 != row2:
                num_diff_records += 1
                if row1 not in sorted_rows2:
                    num_records_file1_not_in_file2 += 1
                if row2 not in sorted_rows1:
                    num_records_file2_not_in_file1 += 1

        # Generate HTML for comparison summary
        summary_html = '<div style="font-size: 15px;">'
        summary_html += f"<p>Total Records in each file: {num_records}</p>"
        summary_html += (
            f"<p>Number of Rows with Differences: {int(num_diff_records)}</p>"
        )
        summary_html += f"<p>Number of records from {file1} differ by from {file2}:  {num_records_file1_not_in_file2}</p>"
        summary_html += f"<p>Number of records from {file2} differ by from {file1}:  {num_records_file2_not_in_file1}</p>"
        summary_html += "</div>\n"

        # Add the summary HTML to the start of the HTML string
        html += summary_html

        # Check if there are any differences in the data and generate HTML table for differences
        if num_diff_records == 0:
            print(
                f"{file1} and {file2} have no differences. Please check the generated HTML for more details."
            )
        elif num_diff_records > 0:
            print(
                f"{file1} and {file2} have differences by {num_diff_records} . Please check the generated HTML for more details."
            )

            diff_rows = []

            # Compare each row from both files and add different rows to diff_rows list
            for row_num, (row1, row2) in enumerate(
                zip(sorted_rows1, sorted_rows2), start=1
            ):
                row_diff = False
                diff_cols = []

                # Check if the values in each column of the row are same
                for col, val1 in row1.items():
                    val2 = row2[col]

                    # If the values in the columns are different, add the column index to diff_cols list
                    if val1 != val2:
                        row_diff = True
                        diff_cols.append(col)

                # If the row has differences, add it to the diff_rows list along with the indices of different columns
                if row_diff:
                    diff_rows.append((row_num, diff_cols, row1, row2))

            # Write the HTML table for differences to the output file
            html += '<table style="font-family: Helvetica, sans-serif; font-size:14px;border-collapse: collapse; width: 100%; border: 1px solid black; border-spacing: 0px;">\n'
            html += '<tr style="background-color: #a0d1dd; border: 1px solid black;"><th colspan="100" style="border: 1px solid black;">Differences</th></tr>'
            html += '<tr style="background-color: #a0d1dd; border: 1px solid black;"><th style="border: 1px solid black;">File</th><th style="border: 1px solid black;">Row Number</th>'

            # Write the headers of CSV files
            for col in header1:
                html += f'<th style="border: 1px solid black;">{col}</th>'

            html += "</tr>\n"

            # Write the rows with differences to the HTML file
            for diff_row in diff_rows:
                html += '<tr style="border: 1px solid black;">'
                html += f'<td style="border: 1px solid black;">{file1}</td>'
                html += f'<td style="border: 1px solid black;">{diff_row[0]}</td>'

                # Write the values of each column of the row
                for col, val in diff_row[2].items():
                    if col in diff_row[1]:
                        html += f'<td style="border: 1px solid black; background-color: #ffcfbf;font-weight: bold;">{val}</td>'
                    else:
                        html += f'<td style="border: 1px solid black;">{val}</td>'

                html += "</tr>\n"

                html += '<tr style="border: 1px solid black;">'
                html += f'<td style="border: 1px solid black;">{file2}</td>'
                html += f'<td style="border: 1px solid black;">{diff_row[0]}</td>'

                # Write the values of each column of the row
                for col, val in diff_row[3].items():
                    if col in diff_row[1]:
                        html += f'<td style="border: 1px solid black; background-color: #ffcfbf;font-weight: bold;">{val}</td>'
                    else:
                        html += f'<td style="border: 1px solid black;">{val}</td>'
                html += "</tr>\n"

            # Close the HTML table
            html += "</table>\n"

        # Write the contents of the entire HTML string to the output file
        outfile.write(html)


compare_csv_files("a.csv", "b.csv", "output.html")
